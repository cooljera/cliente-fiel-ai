import os
import pandas as pd
import joblib
from telegram import Update
from telegram.ext import (
    ApplicationBuilder, CommandHandler, MessageHandler, ContextTypes,
    filters, ConversationHandler
)
from generador_copys import generar_copy
import logging
from dotenv import load_dotenv
from openai import OpenAI  # Cambiado aquí
import openpyxl
from openpyxl.styles import Font, PatternFill
from openpyxl.worksheet.table import Table, TableStyleInfo

# Configuración de logging
logging.basicConfig(
    format="%(asctime)s - %(name)s - %(levelname)s - %(message)s", level=logging.INFO
)

# Cargar variables de entorno
load_dotenv()
client = OpenAI(api_key=os.getenv('OPENAI_API_KEY'))

# Código de prueba para verificar la conexión
try:
    response = client.chat.completions.create(
        model="gpt-3.5-turbo",
        messages=[{"role": "user", "content": "Di hola"}],
        max_tokens=10
    )
    print("✅ OpenAI funcionando correctamente")
    print("Respuesta:", response.choices[0].message.content)
except Exception as e:
    print("❌ Error con OpenAI:", str(e))

# Cargar modelo entrenado
modelo = joblib.load("modelo_segmentador.pkl")

TOKEN = "7708316997:AAFJFk-Rxe554T7LCKB6legPPIptQ_OWcEA"

# Corregir el range para que genere 11 valores (0-10)
(
    PEDIR_CONFIRMACION,
    PEDIR_NOMBRE,
    PEDIR_TIPO,
    PEDIR_DIRECCION,
    PEDIR_WHATSAPP,
    PEDIR_PROMO_AGRESIVA,
    PEDIR_PROMO_FIDELIZACION,
    PEDIR_PROMO_BASICA,
    ESPERAR_ARCHIVO,
    PEDIR_CONFIRMACION_PROMOS,
    ESPERAR_REENVIO
) = range(11)  # Cambiado de 10 a 11

async def start(update: Update, context: ContextTypes.DEFAULT_TYPE):
    await update.message.reply_text(
        "👋 ¡Hola! Soy tu asistente de marketing de Cliente Fiel AI.\n"
        "Voy a ayudarte a segmentar tus clientes y crear promociones personalizadas.\n"
        "Para eso necesito algunos datos de tu empresa.\n\n"
        "¿Te parece bien que comencemos? (Responde 'sí' para continuar)"
    )
    return PEDIR_CONFIRMACION

async def pedir_nombre(update: Update, context: ContextTypes.DEFAULT_TYPE):
    if update.message.text.lower() not in ["sí", "si", "ok", "dale", "claro"]:
        await update.message.reply_text("¡Sin problema! Cuando quieras comenzar, escribe /start.")
        return ConversationHandler.END
    await update.message.reply_text("¿Cómo se llama tu restaurante?")
    return PEDIR_NOMBRE

async def recibir_nombre(update: Update, context: ContextTypes.DEFAULT_TYPE):
    context.user_data['restaurante'] = update.message.text
    await update.message.reply_text("¿Qué tipo de comida ofreces? (ejemplo: comida rápida, carnes, sushi, etc.)")
    return PEDIR_TIPO

async def recibir_tipo(update: Update, context: ContextTypes.DEFAULT_TYPE):
    context.user_data['tipo'] = update.message.text
    await update.message.reply_text("¿Cuál es la dirección de tu restaurante?")
    return PEDIR_DIRECCION

async def recibir_direccion(update: Update, context: ContextTypes.DEFAULT_TYPE):
    context.user_data['direccion'] = update.message.text
    await update.message.reply_text("¿Cuál es el número de WhatsApp para pedidos?")
    return PEDIR_WHATSAPP

async def recibir_whatsapp(update: Update, context: ContextTypes.DEFAULT_TYPE):
    context.user_data['whatsapp'] = update.message.text
    await update.message.reply_text(
        "Trabajaremos con tres tipos de clientes:\n\n"
        "1️⃣ Cliente Normal (anteriormente en riesgo)\n"
        "2️⃣ Cliente VIP (nuestros mejores clientes)\n"
        "3️⃣ Cliente Especial (clientes habituales)\n\n"
        "Ahora necesito que me des la información del tipo de promoción que vas a usar para cada cliente.\n\n"
        "¿Deseas continuar? (Responde 'sí' para continuar)"
    )
    return PEDIR_CONFIRMACION_PROMOS

async def confirmar_promos(update: Update, context: ContextTypes.DEFAULT_TYPE):
    if update.message.text.lower() not in ["sí", "si", "ok", "dale", "claro"]:
        await update.message.reply_text("¡Sin problema! Cuando quieras continuar, escribe /start.")
        return ConversationHandler.END
    
    await update.message.reply_text(
        "Para clientes NORMALES, ¿qué promoción deseas ofrecer?\n"
        "Ejemplo: 2x1 o 50% de descuento"
    )
    return PEDIR_PROMO_AGRESIVA

async def recibir_promo_agresiva(update: Update, context: ContextTypes.DEFAULT_TYPE):
    context.user_data['promo_agresiva'] = update.message.text
    await update.message.reply_text(
        "¿Qué promoción quieres ofrecer a los clientes VIP? (ejemplo: 20% + papas gratis)"
    )
    return PEDIR_PROMO_FIDELIZACION

async def recibir_promo_fidelizacion(update: Update, context: ContextTypes.DEFAULT_TYPE):
    context.user_data['promo_fidelizacion'] = update.message.text
    await update.message.reply_text(
        "¿Qué promoción quieres ofrecer a los clientes comunes? (ejemplo: 10% o bebida gratis)"
    )
    return PEDIR_PROMO_BASICA

async def recibir_promo_basica(update: Update, context: ContextTypes.DEFAULT_TYPE):
    context.user_data['promo_basica'] = update.message.text
    await update.message.reply_text(
        "¡Perfecto! Ahora envíame el archivo Excel de tus clientes.\n\n"
        "El archivo debe tener estas columnas:\n"
        "- NOMBRE DEL CLIENTE\n- NUMERO DE TELEFONO\n- TOTAL DE VISITAS\n- GASTO TOTAL\n- FECHA PRIMERA VISITA\n- FECHA ULTIMA VISITA"
    )
    return ESPERAR_ARCHIVO

async def validar_archivo(df):
    """Valida la estructura y datos del archivo Excel."""
    try:
        # Verificar columnas requeridas
        columnas_requeridas = [
            'NOMBRE DEL CLIENTE', 'NUMERO DE TELEFONO', 'TOTAL DE VISITAS',
            'GASTO TOTAL', 'FECHA PRIMERA VISITA', 'FECHA ULTIMA VISITA'
        ]
        
        columnas_faltantes = [col for col in columnas_requeridas if col not in df.columns]
        if columnas_faltantes:
            return False, f"Faltan columnas: {', '.join(columnas_faltantes)}"

        # Imprimir resumen de datos
        print("\n=== Resumen del archivo ===")
        print(f"Total de registros: {len(df)}")
        print("\nPrimeras 3 filas:")
        print(df.head(3))
        print("\nTipos de datos:")
        print(df.dtypes)
        print("\nEstadísticas básicas:")
        print(df.describe())
        print("========================\n")

        return True, "Archivo válido"

    except Exception as e:
        return False, f"Error validando archivo: {str(e)}"

async def recibir_documento(update: Update, context: ContextTypes.DEFAULT_TYPE):
    try:
        archivo = update.message.document
        file = await archivo.get_file()
        await file.download_to_drive("clientes_entrada.xlsx")
        
        df = pd.read_excel("clientes_entrada.xlsx")
        valido, mensaje = await validar_archivo(df)
        
        if not valido:
            await update.message.reply_text(f"❌ {mensaje}")
            return

        # Procesar datos
        df['GASTO TOTAL'] = df['GASTO TOTAL'].replace(r'[\$,]', '', regex=True).astype(float)
        df['FECHA PRIMERA VISITA'] = pd.to_datetime(df['FECHA PRIMERA VISITA'])
        df['FECHA ULTIMA VISITA'] = pd.to_datetime(df['FECHA ULTIMA VISITA'])
        df['DIAS DESDE ULTIMA VISITA'] = (pd.Timestamp.today() - df['FECHA ULTIMA VISITA']).dt.days

        # Hacer predicción
        X = df[['GASTO TOTAL', 'DIAS DESDE ULTIMA VISITA', 'TOTAL DE VISITAS']]
        df['SEGMENTO'] = modelo.predict(X)
        
        print("Segmentos encontrados:", df['SEGMENTO'].unique())

        # Preparar datos para el Excel - Fix the column name
        df['PRIMER NOMBRE'] = df['NOMBRE DEL CLIENTE'].str.split().str[0]
        
        # Mapear los segmentos numéricos a nombres legibles
        segmento_nombres = {
            '0': 'NORMAL',
            '1': 'VIP',
            '2': 'ESPECIAL'
        }
        df['SEGMENTO'] = df['SEGMENTO'].astype(str).map(segmento_nombres)

        # Generar copys personalizados según el segmento y la promo asignada por el usuario
        for idx, row in df.iterrows():
            segmento = row['SEGMENTO']
            if segmento == 'NORMAL':
                promo = context.user_data.get('promo_agresiva', '')
            elif segmento == 'VIP':
                promo = context.user_data.get('promo_fidelizacion', '')
            elif segmento == 'ESPECIAL':
                promo = context.user_data.get('promo_basica', '')
            else:
                promo = ''
            df.at[idx, 'COPY PERSONALIZADO'] = generar_copy(
                segmento,
                row['PRIMER NOMBRE'],
                context.user_data.get('restaurante', ''),
                context.user_data.get('tipo', ''),
                promo,
                context.user_data.get('direccion', ''),
                context.user_data.get('whatsapp', '')
            )

        # Guardar Excel asegurando al menos una hoja
        nombre_archivo = "clientes_resultado.xlsx"
        with pd.ExcelWriter(nombre_archivo, engine='openpyxl') as writer:
            # Guardar todos los clientes en una hoja
            hoja_general = "TODOS_LOS_CLIENTES"
            columnas = ['PRIMER NOMBRE', 'NUMERO DE TELEFONO', 'SEGMENTO', 'COPY PERSONALIZADO']
            df[columnas].to_excel(writer, sheet_name=hoja_general, index=False)
            
            # Guardar por segmentos
            for segmento in df['SEGMENTO'].unique():
                grupo = df[df['SEGMENTO'] == segmento]
                if not grupo.empty:
                    grupo[columnas].to_excel(writer, sheet_name=segmento, index=False)

        # Formatear cada hoja
        wb = openpyxl.load_workbook(nombre_archivo)
        for hoja in wb.sheetnames:
            ws = wb[hoja]
            # Encabezados en azul y letra grande
            for cell in ws[1]:
                cell.font = Font(color="FFFFFF", bold=True, size=13)
                cell.fill = PatternFill(start_color="1976D2", end_color="1976D2", fill_type="solid")
            # Formato de número para la columna de teléfono
            for row in ws.iter_rows(min_row=2, min_col=2, max_col=2):
                for cell in row:
                    cell.number_format = '0'
            # Aplicar formato de tabla
            rango = f"A1:D{ws.max_row}"
            tabla = Table(displayName=f"Tabla_{hoja.replace(' ', '_')}", ref=rango)
            estilo = TableStyleInfo(name="TableStyleMedium9", showFirstColumn=False,
                                    showLastColumn=False, showRowStripes=True, showColumnStripes=False)
            tabla.tableStyleInfo = estilo
            ws.add_table(tabla)

        wb.save(nombre_archivo)

        await update.message.reply_document(document=open(nombre_archivo, "rb"), filename=nombre_archivo)
        await update.message.reply_text("📄 ¡Archivo procesado con éxito! Los clientes han sido segmentados y sus mensajes personalizados están listos.")

    except Exception as e:
        error_mensaje = str(e)
        error_explicacion = {
            "At least one sheet must be visible": "No se pudo crear ninguna hoja en el Excel. Verifica que los datos sean correctos.",
            "'PRIMER_NOMBRE'": "Error al procesar los nombres de los clientes. Verifica la columna NOMBRE DEL CLIENTE.",
            # Agrega más explicaciones según los errores comunes
        }
        
        explicacion = error_explicacion.get(error_mensaje, "Error desconocido")
        
        await update.message.reply_text(
            f"❌ Se encontró un error al procesar tu archivo:\n"
            f"Error: {error_mensaje}\n"
            f"Explicación: {explicacion}\n\n"
            "Por favor, revisa y corrige el error en tu archivo.\n"
            "Cuando esté listo, escribe 'listo' para enviar el archivo nuevamente."
        )
        return ESPERAR_REENVIO

async def esperar_reenvio(update: Update, context: ContextTypes.DEFAULT_TYPE):
    if update.message.text.lower() == "listo":
        await update.message.reply_text(
            "¡Perfecto! Por favor, envía nuevamente tu archivo Excel.\n"
            "Asegúrate de que contenga todas las columnas requeridas:\n"
            "- NOMBRE DEL CLIENTE\n- NUMERO DE TELEFONO\n- TOTAL DE VISITAS\n"
            "- GASTO TOTAL\n- FECHA PRIMERA VISITA\n- FECHA ULTIMA VISITA"
        )
        return ESPERAR_ARCHIVO
    else:
        await update.message.reply_text(
            "Cuando hayas corregido el archivo, escribe 'listo' para continuar."
        )
        return ESPERAR_REENVIO

async def error_handler(update, context):
    print(f"Error: {context.error}")
    if update and update.message:
        await update.message.reply_text(
            f"❌ Ocurrió un error inesperado: {context.error}"
        )

# Cambiar el ConversationHandler
conv_handler = ConversationHandler(
    entry_points=[CommandHandler("start", start)],  # Quitada la barra "/"
    states={
        PEDIR_CONFIRMACION: [MessageHandler(filters.TEXT & ~filters.COMMAND, pedir_nombre)],
        PEDIR_NOMBRE: [MessageHandler(filters.TEXT & ~filters.COMMAND, recibir_nombre)],
        PEDIR_TIPO: [MessageHandler(filters.TEXT & ~filters.COMMAND, recibir_tipo)],
        PEDIR_DIRECCION: [MessageHandler(filters.TEXT & ~filters.COMMAND, recibir_direccion)],
        PEDIR_WHATSAPP: [MessageHandler(filters.TEXT & ~filters.COMMAND, recibir_whatsapp)],
        PEDIR_CONFIRMACION_PROMOS: [MessageHandler(filters.TEXT & ~filters.COMMAND, confirmar_promos)],
        PEDIR_PROMO_AGRESIVA: [MessageHandler(filters.TEXT & ~filters.COMMAND, recibir_promo_agresiva)],
        PEDIR_PROMO_FIDELIZACION: [MessageHandler(filters.TEXT & ~filters.COMMAND, recibir_promo_fidelizacion)],
        PEDIR_PROMO_BASICA: [MessageHandler(filters.TEXT & ~filters.COMMAND, recibir_promo_basica)],
        ESPERAR_ARCHIVO: [MessageHandler(filters.Document.ALL, recibir_documento)],
        ESPERAR_REENVIO: [MessageHandler(filters.TEXT & ~filters.COMMAND, esperar_reenvio)],
    },
    fallbacks=[CommandHandler("start", start)],  # Quitada la barra "/"
)

app = ApplicationBuilder().token(TOKEN).build()
app.add_handler(conv_handler)
app.add_error_handler(error_handler)

print("🤖 Bot ejecutándose...")
app.run_polling()
